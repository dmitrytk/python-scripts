import openpyxl as op
import pandas as pd
from openpyxl.utils import get_column_letter

from runner import run
from styles import style_basic, style_bold

INPUT_FILE = 'data.xlsx'
OUTPUT_FILE = 'result.xlsx'

months_names = [
    'Январь',
    'Февраль',
    'Март',
    'Апрель',
    'Май',
    'Июнь',
    'Июль',
    'Август',
    'Сентябрь',
    'Октябрь',
    'Ноябрь',
    'Декабрь',
]


def main():
    df = pd.read_excel(INPUT_FILE)
    wb = op.Workbook()
    ws = wb.active
    ws.append(['Месяц'] + ['Параметр'] + [i for i in range(1, 32)])

    wells = list(dict.fromkeys(df.well))

    # Iterate over well
    for well in wells:
        print(well)
        ws.append([well])
        last_row = ws.max_row
        ws.merge_cells(start_row=last_row, start_column=1,
                       end_row=last_row, end_column=33)

        well_df = df[df.well == well]
        years = list(dict.fromkeys(well_df.date.dt.year))
        # Iterate over year
        for year in years:
            year_df = well_df[well_df.date.dt.year == year]
            months = list(dict.fromkeys(year_df.date.dt.month))
            for month in months:
                month_df = year_df[year_df.date.dt.month == month]

                # Blank rows
                q = [None for i in range(31)]
                dynamic = [None for i in range(31)]
                static = [None for i in range(31)]
                # Iterate over month data
                for _, row in month_df.iterrows():
                    day = row.date.day
                    q[day - 1] = round(row.rate, 1)
                    dynamic[day - 1] = round(row.dynamic, 1)
                    static[day - 1] = round(row.static, 1)

                # Write rows to sheet
                ws.append([f'{months_names[month - 1]} {year}'] +
                          ['Q, м3/сут'] + q)
                ws.append([None] + ['Нд, м'] + dynamic)
                ws.append([None] + ['Нст, м'] + static)
                last_row = ws.max_row
                ws.merge_cells(start_row=last_row - 2, start_column=1,
                               end_row=last_row, end_column=1)

    # Apply styles
    for row in ws.iter_rows():
        for cell in row:
            cell.style = style_basic

        if row[0].value is not None and row[1].value is None:
            row[0].style = style_bold

    for i in range(2, 33):
        ws.column_dimensions[get_column_letter(i + 1)].width = 6

    wb.save(OUTPUT_FILE)


# -------------------MAIN----------------------#
if __name__ == '__main__':
    run(main)
