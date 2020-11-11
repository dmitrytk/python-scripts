#!/usr/bin/env python
"""
Parse Salym Petroleum rate files
"""

import os

import openpyxl as op

from runner import run


def main():
    xlsx_files = [i for i in os.listdir(
    ) if 'xlsx' in i and '~' not in i and 'result' not in i]

    wb_result = op.Workbook()
    ws_result = wb_result.active
    ws_result.append(['well', 'date', 'rate', 'dynamic', 'static'])

    for file in xlsx_files:
        print(file)
        wb = op.load_workbook(file)

        for sheet in wb.worksheets:
            if sheet['B1'].value is None:
                continue
            else:
                well_name = sheet['B1'].value.replace("US-", "")
                for row in sheet.iter_rows(min_row=9):
                    if row[1].value is None:
                        break
                    row_data = [well_name, row[1].value, row[4].value,
                                row[5].value, row[6].value]
                    ws_result.append(row_data)

    wb_result.save('result.xlsx')


# -------------------MAIN----------------------#
if __name__ == '__main__':
    run(main)
