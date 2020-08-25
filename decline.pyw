#!/usr/bin/env python
'''
Simple tkinter program to calculate water table decline.
Draw table, map and grid!
Input file: ./templates/sub_template.xlsx
'''

from tkinter import *
from tkinter import messagebox
import tkinter.ttk as ttk
import time
import threading
from tkinter import filedialog
import os
import matplotlib.pyplot as plt
import numpy as np
from scipy.special import expi
import math
import openpyxl as op
import matplotlib.pyplot as plt
import datetime
import traceback


# CLASSES
class Data:
    def __init__(self):
        self.km = 0
        self.a = 0
        self.years = []
        self.dates = []
        self.wells = []
        self.dist = []


class Grid:
    def __init__(self):
        self.XX = []
        self.YY = []
        self.X = 0
        self.Y = 0
        self.Z = 0
        self.calculated = False


class Well:
    def __init__(self, name, x, y, r):
        self.name = name
        self.x = x
        self.y = y
        self.r = r
        self.debts = []
        self.S = []

# FUNCTIONS


def log(exc):
    '''Log errors to file'''
    now = datetime.datetime.now()
    try:
        with open(f'error_log_{now.hour}-{now.minute}.txt', 'w') as file:
            file.write(traceback.format_exc())
    except Exception as e:
        print(e)


def distance(well, x, y):
    '''Calculate distance between well and point'''
    return ((well.x - x)**2 + (well.y - y)**2)**0.5


def load_excel_file(file_name):
    '''Load excel data file'''
    wb = op.load_workbook(file_name)
    ws = wb['DATA']
    data = Data()

    start_date = ws[5][4].value
    last_cell = len(ws[5])
    for index, cell in enumerate(ws[5][4:]):
        if cell.value == None or cell.value == '':
            last_cell = index

    for cell in ws[5][4:last_cell]:
        if cell.value == '' or cell.value == None:
            break
        data.dates.append((cell.value - start_date).days)
        data.years.append(cell.value)

    data.km = ws[2][0].value
    data.a = ws[2][1].value

    for row in ws.iter_rows(min_row=6):
        if row[0].value == '':
            break
        well = Well(row[0].value, row[1].value, row[2].value, row[3].value)
        well.debts = [
            int(cell.value) for cell in row[4:last_cell]
            if cell.value != None and cell.value != ''
        ]
        data.wells.append(well)

    for well in data.wells:
        well.debts[0] = 0

    for i, well in enumerate(data.wells):
        arr = []
        for j, other_well in enumerate(data.wells):
            if i == j:
                arr.append(other_well.r)
            else:
                arr.append(distance(well, other_well.x, other_well.y))
        data.dist.append(arr)

    return data


def calculate_table(data):
    '''Calculate decline table'''
    wb_out = op.Workbook()
    ws_out = wb_out.active
    first_row = ['WELL', 'X', 'Y', 'R']

    for date in data.years:
        first_row.append(date.date())
    ws_out.append(first_row)

    for i, final_year in enumerate(data.dates):
        for j, well in enumerate(data.wells):
            SUM = 0
            for k in range(1, i + 1):
                for l, other_well in enumerate(data.wells):
                    R = data.dist[j][l]
                    Q = other_well.debts[k] - other_well.debts[k - 1]
                    T = data.dates[i] - data.dates[k - 1]
                    arg = -1 * R * R / 4 / data.a / T / 100000
                    SUM += Q * expi(arg)
            well.S.append(-1 / 4 / 3.14 / data.km * SUM)
        print(f'Done {i} year')

    for well in data.wells:
        row = [well.name, well.x, well.y, well.r]
        for s in well.S:
            row.append(s)
        ws_out.append(row)

    wb_out.save('RESULT.xlsx')
    os.system('start ""  RESULT.xlsx"')


def calculate_grid(data, grid, margin, step, pb):
    '''Calculate water table decline 2d grid'''
    grid = Grid()
    grid.XX = [well.x for well in data.wells]
    grid.YY = [well.y for well in data.wells]

    min_x = math.floor((min(grid.XX) - margin) / 1000) * 1000
    max_x = math.ceil((max(grid.XX) + margin) / 1000) * 1000
    min_y = math.floor((min(grid.YY) - margin) / 1000) * 1000
    max_y = math.ceil((max(grid.YY) + margin) / 1000) * 1000

    x_list = np.arange(min_x, max_x + step, step)
    y_list = np.arange(min_y, max_y + step, step)

    grid.X, grid.Y = np.meshgrid(x_list, y_list)
    grid.Z = np.zeros((len(grid.X), len(grid.X[1])))

    num_points = len(grid.Z) * len(grid.Z[0])
    percent = round(num_points / 100) * 1
    counter = 0

    for i, val_x in enumerate(grid.X):
        for j, val_y in enumerate(grid.X[i]):
            SUM = 0
            for well in data.wells:
                R = distance(well, grid.X[i][j], grid.Y[i][j])
                for k, date in enumerate(data.dates):
                    if k > 0:
                        Q = well.debts[k] - well.debts[k - 1]
                        T = data.dates[-1] - data.dates[k - 1]
                        arg = -1 * R * R / 4 / data.a / T / 100000
                        SUM += Q * expi(arg)

            temp_S = -1 / 4 / 3.14 / data.km * SUM
            grid.Z[i][j] = round(temp_S, 2)
            counter += 1
            if counter % percent == 0:
                pb['value'] += 1
    pb.grid_remove()
    return grid


def write_grid(grid):
    '''Write grid data to file'''
    grd = 'DSAA\n'
    grd += str(len(grid.X[0])) + ' ' + str(len(grid.X)) + '\n'
    grd += str(grid.X[0][0]) + ' ' + str(grid.X[0][-1]) + '\n'
    grd += str(grid.Y[0][0]) + ' ' + str(grid.Y[-1][0]) + '\n'

    counter = 0
    grd += str(np.amax(grid.Z)) + ' ' + str(np.amin(grid.Z)) + ' \n'

    for i, val_x in enumerate(grid.X):
        for j, val_y in enumerate(grid.X[i]):
            grd += str(grid.Z[i][j]) + ' '
            counter += 1
            if counter % 10 == 0:
                grd += '\n'
    with open('RESULT.grd', 'w') as file:
        file.write(grd)
    os.system('start "" notepad RESULT.grd')


def plot_grid(data, grid):
    fig = plt.figure(figsize=(12, 10))
    left, bottom, width, height = 0.1, 0.1, 0.8, 0.8
    ax = fig.add_axes([left, bottom, width, height])
    ax.plot(grid.XX, grid.YY, 'ko', ms=3)

    ax.axis('equal')

    cp = plt.contourf(grid.X, grid.Y, grid.Z)
    plt.colorbar(cp)

    ax.set_title('Понижения')
    ax.set_xlabel('X')
    ax.set_ylabel('Y')

    for well in data.wells:
        ax.annotate(well.name, (well.x, well.y))

    plt.savefig("RESULT.png")
    os.system('start ""  RESULT.png"')


# GUI
root = Tk()
root.geometry("+800+200")
root.resizable(False, False)
root.title("SUB")
root.DATA = 0
root.GRID = 0
root.GRID_CALCULATED = False

step_label = Label(text="Шаг сетки, м")
margin_label = Label(text="Запас по X и Y, м")
step_label.grid(row=0, column=0, sticky="w")
margin_label.grid(row=1, column=0, sticky="w")

step_entry = Entry()
margin_entry = Entry()
step_entry.grid(row=0, column=1, padx=2, pady=2)
margin_entry.grid(row=1, column=1, padx=2, pady=2)
step_entry.insert(END, '100')
margin_entry.insert(END, '3000')
pb = ttk.Progressbar(root, mode="determinate")


def show_progressbar():
    pb['value'] = 0
    pb.grid(
        row=6,
        column=0,
        columnspan=2,
        rowspan=4,
        padx=2,
        pady=2,
        sticky=E + W + S + N)


def hide_progressbar():
    pb.grid_remove()


def load_file():
    filename = filedialog.askopenfilename(
        initialdir=os.getcwd(),
        title="Select file",
        filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
    if len(filename) > 0:
        try:
            root.DATA = load_excel_file(filename)
            root.title("Correct file")
            calc_button['state'] = 'normal'
            root.GRID_CALCULATED = False
        except Exception as e:
            log(e)
            calc_button['state'] = 'disabled'
            root.title("Incorrect file")


# RADIO BUTTONS
table_var = BooleanVar()
table_var.set(1)
table_btn = Checkbutton(
    text="Таблица", variable=table_var, onvalue=1, offvalue=0)
table_btn.grid(row=2, column=0, padx=2, pady=2, sticky=W)

grid_var = BooleanVar()
grid_var.set(0)
grid_btn = Checkbutton(text="Грид", variable=grid_var, onvalue=1, offvalue=0)
grid_btn.grid(row=3, column=0, padx=2, pady=2, sticky=W)

image_var = BooleanVar()
image_var.set(0)
image_btn = Checkbutton(
    text="Картинка", variable=image_var, onvalue=1, offvalue=0)
image_btn.grid(row=4, column=0, padx=2, pady=2, sticky=W)


def calc():
    """Calculate table and grid"""
    try:
        if table_var.get() == 1:
            root.title('Processing')
            calculate_table(root.DATA)
            root.title('Done!')
        if grid_var.get() == 1 or image_var.get() == 1:
            root.step = int(step_entry.get())
            root.margin = int(margin_entry.get())
            if root.GRID_CALCULATED != True:
                root.title('Processing')
                show_progressbar()
                root.GRID = calculate_grid(root.DATA, root.GRID, root.margin,
                                           root.step, pb)
                root.GRID_CALCULATED = True
                root.title('Done!')
            if grid_var.get() == 1:
                write_grid(root.GRID)
            if image_var.get() == 1:
                plot_grid(root.DATA, root.GRID)
    except Exception as e:
        log(e)


def calculate():
    threading.Thread(target=calc).start()


# BUTTONS
load_button = Button(text="Загрузить файл", command=load_file)
load_button.grid(row=5, column=0, padx=2, pady=2, sticky="e")

calc_button = Button(text="Расчет", command=calculate)
calc_button.grid(row=5, column=1, padx=2, pady=2, sticky=E + W)
calc_button['state'] = 'disabled'
# endregion GUI

root.mainloop()
