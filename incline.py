#create separate .inc file for each well
#Excel file template
#WELL    MD  INCL    AZIM
#551 10  0.15    0
#551 20  0.15    0
#551 30  0.27    0

import openpyxl as op
import os
import glob
import time

os.system("cls")


class Well:
    def __init__(self, name):
        self.name = name
        self.md = []
        self.inc = []
        self.azim = []


def main():
    wells = []

    wb = op.load_workbook("INCL.xlsx")
    ws = wb.active

    wells.append(Well(ws[2][0].value))
    wells[-1].md.append(ws[2][1].value)
    wells[-1].inc.append(ws[2][2].value)
    wells[-1].azim.append(ws[2][3].value)

    for row in ws.iter_rows(min_row=3):
        if wells[-1].name == row[0].value:
            wells[-1].md.append(row[1].value)
            wells[-1].inc.append(row[2].value)
            wells[-1].azim.append(row[3].value)
        else:
            wells.append(Well(row[0].value))
            wells[-1].md.append(row[1].value)
            wells[-1].inc.append(row[2].value)
            wells[-1].azim.append(row[3].value)

    print(len(wells))

    try:
        os.mkdir("inc")
    except Exception as e:
        pass

    for well in wells:
        txt = "MD\tINC\tAZIM\n"
        for i in range(len(well.md)):
            txt += str(well.md[i]) + "\t" + str(well.inc[i]) + "\t" + str(
                well.azim[i]) + "\n"

        with open(f"inc/{well.name}.inc", "w") as f:
            f.write(txt)

    print(f"Done in {time.perf_counter()} sec")


if __name__ == "__main__":
    try:
        main()
        input("Press enter to exit")
    except Exception as e:
        print(e)
        input("Press enter to exit")
