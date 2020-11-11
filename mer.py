#!/usr/bin/env python
"""
Parse XMAO mer files
"""

import datetime
import os
import re

import openpyxl as op

from runner import run

months = {
    'январь': 1,
    'февраль': 2,
    'март': 3,
    'апрель': 4,
    'май': 5,
    'июнь': 6,
    'июль': 7,
    'август': 8,
    'сентябрь': 9,
    'октябрь': 10,
    'ноябрь': 11,
    'декабрь': 12,
}


def get_date(cell_value):
    date_row = re.split(' +', cell_value)
    m = months[date_row[2].lower()]
    y = int(date_row[3])
    dt = datetime.date(y, m, 1)
    return dt.strftime('%d.%m.%Y')


def main():
    xlsx_files = [i for i in os.listdir(
    ) if 'xlsx' in i and '~' not in i and 'result' not in i]

    wb_result = op.Workbook()
    ws_result = wb_result.active

    for file in xlsx_files:
        print(file)
        wb = op.load_workbook(file)
        ws = wb.active

        dt = get_date(ws['A2'].value)

        field = 'UNKNOWN'

        for row in ws.iter_rows(min_row=8):
            cell_0 = row[0].value
            cell_1 = row[1].value
            if cell_0 is None and cell_1 is None:
                break
            if cell_0 is not None and 'местор' in cell_0.lower():
                field = re.split(' +', cell_0)[2]
            if cell_1 is not None:
                row_data = [i.value for i in row[1:]]
                ws_result.append([field] + [dt] + row_data)

    wb_result.save('result.xlsx')


# -------------------MAIN----------------------#
if __name__ == '__main__':
    run(main)
