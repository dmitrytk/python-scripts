'''
Create .rb9 file from layers table.
Формат входного файла template.xlsx:
Sheet 1: 'LAYERS'
WELL	ZKA	ZPA
1W	-839.3	-840.5
1W	-840.5	-841.9

Sheet 2: 'WELLS'
WELL	ALT	X	Y
18P	95.8	3543229	726515
1W	107.7	3543302	722886

Sheet 3: 'ZONES'
WELL	ZONE	ZKA	ZPA
18P	УВАТСКАЯ	-844.2	-1583.6
23R	УВАТСКАЯ	-846.8	-1579.2
'''

import os
import time
import openpyxl as op

class Well:
    def __init__(self, name, alt, x, y):
        self.name = name
        self.alt = alt
        self.x = x
        self.y = y
        self.zk = []
        self.zp = []

input_file = "RB9_template.xlsx"
output_file = "RB9_result.xlsx"


def main():
    wb = op.load_workbook(input_file)
    wbOut = op.Workbook()

    wsLayers = wb["LAYERS"]
    wsWells = wb["WELLS"]
    wsZones = wb["ZONES"]
    wells = []

    for row in wsWells.iter_rows(min_row=2):
        wells.append(
            Well(
                str(row[0].value),
                row[1].value,
                row[2].value,
                row[3].value,
            ))

    for well in wells:
        for row in wsLayers.iter_rows(min_row=2):
            if str(row[0].value) == well.name:
                well.zk.append(-1 * row[1].value + well.alt)
                well.zp.append(-1 * row[2].value + well.alt)

    rb9 = ""
    for well in wells:
        rb9 += '"' + str(well.name) + '"\n'
        rb9 += str(round(well.zk[0], 1)) + " " + str(round(
            well.zp[-1], 1)) + " 0.0 " + str(len(well.zk)) + "\n"
        for i in range(len(well.zk)):
            rb9 += str(round(well.zk[i], 1)) + " " + str(
                round((well.zp[i] - well.zk[i]), 1)) + " 0.7\n"

    with open("rigis.rb9", "w") as file:
        file.write(rb9)

    wsOut = wbOut.active
    wsOut.title = "ZONES"
    wsOut.append(["WELL", "ZONE", "ZK", "ZP"])
    for well in wells:
        for row in wsZones.iter_rows(min_row=2):
            if well.name == str(row[0].value):
                r = [row[0].value, row[1].value]
                r.append(-1 * row[2].value + well.alt)
                r.append(-1 * row[3].value + well.alt)
                wsOut.append(r)

    wbOut.save(output_file)
    os.system("start " + output_file)

    print(f"Done in {time.perf_counter()} sec")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(e)
        input("Press Enter to exit")